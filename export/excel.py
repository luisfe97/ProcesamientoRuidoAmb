from openpyxl import load_workbook
from openpyxl.styles import Font, Border, PatternFill, Alignment, Protection
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from utils.file_utils import round_dataframe, is_merged_cell

def export_to_template(TablaProcesada, diurno_grouped, nocturno_grouped, resumen_diurno, resumen_nocturno, dia_noche, template_path, output_path, Estacion):
    """
    Exporta los resultados a una plantilla Excel
    
    Args:
        TablaProcesada: DataFrame con tabla de datos procesados
        diurno_grouped: DataFrame con datos diurnos agrupados
        nocturno_grouped: DataFrame con datos nocturnos agrupados
        resumen_diurno: DataFrame con resumen diurno
        resumen_nocturno: DataFrame con resumen nocturno
        dia_noche: DataFrame con datos combinados de día y noche
        template_path: Ruta a la plantilla Excel
        output_path: Ruta donde guardar el archivo de salida
        Estacion: Nombre de la estación
    """
    wb = load_workbook(template_path)
    ws = wb["Hoja1"]  # Seleccionar la hoja específica
    
    # Agregar la variable Estacion en la celda B1
    ws["B1"] = Estacion
    ws["B1"].alignment = Alignment(horizontal='center', vertical='center')
    ws["B1"].font = Font(bold=True)
    ws["B1"].font = Font(color="FFFFFF") 

    # Redondear todos los datos a 2 decimales
    datasets = [TablaProcesada, diurno_grouped, nocturno_grouped, resumen_diurno, resumen_nocturno, dia_noche]
    datasets = [round_dataframe(df) for df in datasets]

    start_columns = [1, 9, 29, 49, 67, 85]  # Columnas específicas en la plantilla

    # Definir colores
    colors = {
        "pasa": PatternFill(start_color="C6EFC0", end_color="C6EFC0", fill_type="solid"),  # Verde
        "no pasa": PatternFill(start_color="FFC7C0", end_color="FFC7C0", fill_type="solid"),  # Rojo
        "pasa condicional": PatternFill(start_color="FFD490", end_color="FFD490", fill_type="solid")  # Amarillo
    }

    column_widths = {}  # Diccionario para almacenar los anchos máximos de cada columna

    for df, col_start in zip(datasets, start_columns):
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=10):  # No incluir nombres de columnas
            for c_idx, value in enumerate(row, start=col_start):
                if not is_merged_cell(ws, r_idx, c_idx):  # Evitar celdas combinadas
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Aplicar alineación centrada
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Aplicar color de celda si la celda tiene color
                    if isinstance(value, str) and value.lower() in colors:
                        cell.fill = colors[value.lower()]  # Aplicar color según el valor

                    # Ajustar el ancho de la columna basado en la longitud del contenido con un margen extra
                    column_letter = get_column_letter(c_idx)
                    column_widths[column_letter] = max(column_widths.get(column_letter, 0), len(str(value)) + 7)

    # Aplicar los anchos de columna con margen extra
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(output_path)
    print(f"Archivo '{output_path}' guardado con éxito.")