#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Procesamiento de datos de ruido para múltiples estaciones (EMRI).
Este script procesa archivos Excel que contienen datos de ruido para diferentes estaciones,
extrae la información relevante y genera reportes formateados.
"""

import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy


def procesar_excel_simple(ruta_archivo, ruta_guardado=None):
    """
    Procesa un archivo Excel con datos de ruido de múltiples estaciones y genera reportes formateados.
    
    Args:
        ruta_archivo (str): Ruta al archivo Excel a procesar
        ruta_guardado (str, optional): Ruta donde guardar los resultados. Por defecto usa directorio actual.
    
    Returns:
        dict: Diccionario con los DataFrames generados
    """
    # Si no se especifica ruta de guardado, usar el directorio actual
    if ruta_guardado is None:
        ruta_guardado = os.getcwd()
    
    # Asegurar que el directorio de guardado exista
    if not os.path.exists(ruta_guardado):
        os.makedirs(ruta_guardado)
    
    # Cargar el Excel
    xl = pd.ExcelFile(ruta_archivo)
    
    # Filtrar las hojas que empiezan con "PTO"
    hojas_pto = sorted([hoja for hoja in xl.sheet_names if hoja.startswith('PTO')])
    
    print(f"Hojas PTO encontradas: {hojas_pto}")
    
    # Nombres de columnas para los DataFrames en el orden correcto
    nombres_columnas = ["LASeqk", "LAIeqk", "LRASeqk", "sk2", "sk", "TU", "k", "U", "E", "w", "AU", "z", "Rp*=Pc", "Rc", "Declaración"]
    
    # Índices de columnas fijos
    col_ay_idx = 50  # Índice para AY
    col_bm_idx = 64  # Índice para BM
    col_bq_idx = 68  # Índice para BQ
    col_ce_idx = 82  # Índice para CE
    
    # Listas para almacenar los datos
    datos_estaciones = []  # Lista para almacenar tuplas (nombre_estacion, num_estacion, datos_diurnos, datos_nocturnos)
    
    # Procesar cada hoja en paralelo
    for hoja in hojas_pto:
        df = pd.read_excel(xl, sheet_name=hoja)
        
        # Buscar EMRI en cualquier lugar
        nombre_estacion = None
        num_estacion = None
        
        # Buscar en todas las filas y columnas (primeras filas para mayor eficiencia)
        for fila in range(min(10, len(df))):
            for col in range(min(10, df.shape[1])):
                if pd.notna(df.iloc[fila, col]):
                    valor = str(df.iloc[fila, col])
                    if "EMRI" in valor:
                        nombre_estacion = valor.strip()
                        # Extraer el número que sigue a EMRI_
                        match = re.search(r'EMRI_(\d+)', valor)
                        if match:
                            num_estacion = int(match.group(1))
                            print(f"Encontrado {nombre_estacion} en columna con número {num_estacion}")
                        break
            if nombre_estacion:
                break
        
        # Revisar las columnas también
        if nombre_estacion is None:
            for i, col_name in enumerate(df.columns):
                if "EMRI" in str(col_name):
                    nombre_estacion = str(col_name)
                    match = re.search(r'EMRI_(\d+)', nombre_estacion)
                    if match:
                        num_estacion = int(match.group(1))
                        print(f"Encontrado {nombre_estacion} en columna con número {num_estacion}")
                    break
        
        # Si no se encuentra número, usar índice de la hoja
        if nombre_estacion is None:
            nombre_estacion = f"EMRI_{hoja[-1]}"
            num_estacion = int(hoja[-1]) if hoja[-1].isdigit() else 0
            print(f"Usando nombre genérico: {nombre_estacion}")
        elif num_estacion is None:
            # Si se encontró EMRI pero sin número, intentar extraerlo de la hoja
            num_estacion = int(hoja[-1]) if hoja[-1].isdigit() else 0
            print(f"Número no encontrado en EMRI, usando número de hoja: {num_estacion}")
        
        # Extraer datos para filas 9, 10, 11
        datos_diurnos = {}
        datos_nocturnos = {}
        
        for fila, idx in zip([9, 10, 11], [8, 9, 10]):
            if idx < len(df):
                # Datos diurnos (AY-BM)
                if col_ay_idx < df.shape[1] and col_bm_idx < df.shape[1]:
                    valores_diurnos = df.iloc[idx, col_ay_idx:col_bm_idx+1].values.tolist()
                    valores_diurnos = [None if pd.isna(v) else v for v in valores_diurnos]
                    # Ajustar longitud a nombres_columnas
                    if len(valores_diurnos) < len(nombres_columnas):
                        valores_diurnos += [None] * (len(nombres_columnas) - len(valores_diurnos))
                    else:
                        valores_diurnos = valores_diurnos[:len(nombres_columnas)]
                    datos_diurnos[fila] = valores_diurnos
                else:
                    datos_diurnos[fila] = [None] * len(nombres_columnas)
                
                # Datos nocturnos (BQ-CE)
                if col_bq_idx < df.shape[1] and col_ce_idx < df.shape[1]:
                    valores_nocturnos = df.iloc[idx, col_bq_idx:col_ce_idx+1].values.tolist()
                    valores_nocturnos = [None if pd.isna(v) else v for v in valores_nocturnos]
                    # Ajustar longitud a nombres_columnas
                    if len(valores_nocturnos) < len(nombres_columnas):
                        valores_nocturnos += [None] * (len(nombres_columnas) - len(valores_nocturnos))
                    else:
                        valores_nocturnos = valores_nocturnos[:len(nombres_columnas)]
                    datos_nocturnos[fila] = valores_nocturnos
                else:
                    datos_nocturnos[fila] = [None] * len(nombres_columnas)
            else:
                datos_diurnos[fila] = [None] * len(nombres_columnas)
                datos_nocturnos[fila] = [None] * len(nombres_columnas)
            
            print(f"Fila {fila} extraída para {nombre_estacion}")
        
        # Guardar los datos de esta estación
        datos_estaciones.append((nombre_estacion, num_estacion, datos_diurnos, datos_nocturnos))
    
    # Ordenar las estaciones por el número extraído
    datos_estaciones.sort(key=lambda x: x[1])
    
    print("\nEstaciones ordenadas:")
    for nombre, num, _, _ in datos_estaciones:
        print(f"{nombre} (número {num})")
    
    # Ahora crear los DataFrames con las estaciones ordenadas
    dataframes = {}
    
    for fila in [9, 10, 11]:
        # Crear listas para los DataFrames
        nombres = [estacion[0] for estacion in datos_estaciones]
        valores_diurnos = [estacion[2][fila] for estacion in datos_estaciones]
        valores_nocturnos = [estacion[3][fila] for estacion in datos_estaciones]
        
        # DataFrame diurno
        df_diurno = pd.DataFrame(valores_diurnos, columns=nombres_columnas)
        df_diurno.insert(0, 'Nombre', nombres)
        nombre_diurno = f'diurno_fila{fila}'
        dataframes[nombre_diurno] = df_diurno
        

        # DataFrame nocturno
        df_nocturno = pd.DataFrame(valores_nocturnos, columns=nombres_columnas)
        df_nocturno.insert(0, 'Nombre', nombres)
        nombre_nocturno = f'nocturno_fila{fila}'
        dataframes[nombre_nocturno] = df_nocturno    
    # Crear archivo Excel con formato especial
    crear_excel_formato_especial(dataframes, ruta_guardado)
    
    # Verificar dimensiones de cada DataFrame
    for nombre, df in dataframes.items():
        print(f"DataFrame {nombre}: {df.shape[0]} filas x {df.shape[1]} columnas")
        
    return dataframes


def crear_excel_formato_especial(dataframes, ruta_guardado):
    """
    Crea un archivo Excel con formato especial para los DataFrames.
    
    Args:
        dataframes (dict): Diccionario con los DataFrames a guardar
        ruta_guardado (str): Ruta donde guardar el archivo Excel
    """
    ruta_excel_nuevo = os.path.join(ruta_guardado, "RUIDO TOTAL.xlsx")
    
    with pd.ExcelWriter(ruta_excel_nuevo, engine='xlsxwriter') as writer:
        workbook = writer.book
        worksheet = workbook.add_worksheet('RUIDO TOTAL')
        
        # Definir formatos
        formato_titulo_diurno = workbook.add_format({
            'bold': True, 'font_size': 12, 'align': 'center', 'valign': 'vcenter',
            'bg_color': '#FFB47F', 'border': 1  # Naranja (255,180,127)
        })
        
        formato_titulo_nocturno = workbook.add_format({
            'bold': True, 'font_size': 12, 'align': 'center', 'valign': 'vcenter',
            'bg_color': '#202D4D', 'font_color': 'white', 'border': 1  # Azul oscuro (32,45,77)
        })
        
        formato_nombre_columna = workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'bg_color': '#B8AFAC', 'border': 1  # Gris (184,175,172)
        })
        
        formato_encabezado_diurno = workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'bg_color': '#FFB47F', 'border': 1  # Naranja (255,180,127)
        })
        
        formato_encabezado_nocturno = workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'bg_color': '#202D4D', 'font_color': 'white', 'border': 1  # Azul oscuro (32,45,77)
        })
        
        formato_datos = workbook.add_format({
            'align': 'center', 'border': 1
        })
        
        # Formatos para declaraciones
        formato_pasa = workbook.add_format({
            'align': 'center', 'border': 1, 'bg_color': '#C6EFC0', 'font_color': 'green'
        })
        
        formato_no_pasa = workbook.add_format({
            'align': 'center', 'border': 1, 'bg_color': '#FFC7C0', 'font_color': 'red'
        })
        
        formato_pasa_condicional = workbook.add_format({
            'align': 'center', 'border': 1, 'bg_color': '#FFD490', 'font_color': '#B8860B'
        })
        
        # Formato para EMRI (nombres azules)
        formato_emri = workbook.add_format({
            'align': 'left', 'border': 1, 'font_color': 'blue'
        })
        
        # Formato para números
        formato_numero = workbook.add_format({
            'align': 'center', 'border': 1, 'num_format': '0.0'
        })
        
        # Títulos para los dataframes
        titulos_diurnos = ["DIURNO DOMINICAL", "DIURNO ORDINARIO", "DIURNO GLOBAL"]
        titulos_nocturnos = ["NOCTURNO DOMINICAL", "NOCTURNO ORDINARIO", "NOCTURNO GLOBAL"]
        
        # Calcular la fila inicial para los nocturnos
        max_filas_diurno = max([dataframes[f'diurno_fila{fila}'].shape[0] for fila in [9, 10, 11]])
        fila_inicio_nocturnos = max_filas_diurno + 2 + 5  # +2 por encabezados, +5 por espacio
        
        # Configuración de columnas - ancho adecuado para cada tipo de dato
        worksheet.set_column('A:A', 16)  # Columna Nombre
        
        # Diccionario para rastrear el ancho máximo necesario para cada columna
        anchos_columnas = {}
        
        # Inicializar anchos mínimos basados en longitud de encabezados
        for j, col_name in enumerate(dataframes['diurno_fila9'].columns[1:]):
            anchos_columnas[j] = len(str(col_name)) + 2  # +2 para margen
        
        # Escribir los dataframes diurnos en horizontal
        col_offset = 0
        for i, fila in enumerate([9, 10, 11]):
            df_diurno = dataframes[f'diurno_fila{fila}']
            
            # Determinar qué columnas escribir y calcular el ancho del título
            if i == 0:  # Para el primer dataframe, incluir columna 'Nombre'
                worksheet.write(1, 0, "Nombre", formato_nombre_columna)
                ancho_titulo = len(df_diurno.columns) 
                col_offset_data = 1
                
                # Calcular ancho máximo de la columna Nombre
                max_nombre_ancho = len("Nombre") + 2
                for nombre in df_diurno['Nombre']:
                    max_nombre_ancho = max(max_nombre_ancho, len(str(nombre)) + 2)
                worksheet.set_column(0, 0, max_nombre_ancho)
            else:  # Para los otros dataframes, omitir columna 'Nombre'
                ancho_titulo = len(df_diurno.columns) - 1
                col_offset_data = 0
            
            # Escribir el título combinando celdas
            worksheet.merge_range(0, col_offset, 0, col_offset + ancho_titulo - 1, 
                                 titulos_diurnos[i], formato_titulo_diurno)
            
            # Escribir los encabezados
            nombres_columnas = df_diurno.columns[1:]  # Excluir 'Nombre'
            for j, col_name in enumerate(nombres_columnas):
                worksheet.write(1, col_offset + col_offset_data + j, col_name, formato_encabezado_diurno)
            
            # Escribir los datos
            for row_idx in range(df_diurno.shape[0]):
                # Escribir el nombre en la primera columna solo para el primer dataframe
                if i == 0:
                    nombre_estacion = df_diurno.iloc[row_idx, 0]
                    worksheet.write(row_idx + 2, 0, nombre_estacion, formato_emri)
                
                # Escribir los valores de las columnas de datos
                for j, col_idx in enumerate(range(1, df_diurno.shape[1])):
                    valor = df_diurno.iloc[row_idx, col_idx]
                    col_real = col_offset + col_offset_data + j
                    
                    # Si la columna es "Declaración", aplicar formato según el valor
                    if df_diurno.columns[col_idx] == "Declaración":
                        if pd.notna(valor):
                            valor_lower = str(valor).lower()
                            if valor_lower == "pasa":
                                formato = formato_pasa
                            elif valor_lower == "no pasa":
                                formato = formato_no_pasa
                            elif "condicional" in valor_lower:
                                formato = formato_pasa_condicional
                            else:
                                formato = formato_datos
                            
                            worksheet.write(row_idx + 2, col_real, valor, formato)
                            # Actualizar ancho máximo
                            ancho_actual = anchos_columnas.get(j, 0)
                            anchos_columnas[j] = max(ancho_actual, len(str(valor)) + 2)
                    else:
                        # Para columnas numéricas
                        if pd.notna(valor):
                            if isinstance(valor, (int, float)):
                                worksheet.write_number(row_idx + 2, col_real, valor, formato_numero)
                            else:
                                worksheet.write(row_idx + 2, col_real, valor, formato_datos)
                            # Actualizar ancho máximo
                            ancho_actual = anchos_columnas.get(j, 0)
                            anchos_columnas[j] = max(ancho_actual, len(str(valor)) + 2)
            
            # Actualizar el desplazamiento para el siguiente dataframe
            col_offset += ancho_titulo + 1  # +1 para dejar espacio entre dataframes
            
            # Aplicar anchos de columna
            for j, ancho in anchos_columnas.items():
                col_num = col_offset - ancho_titulo + j
                worksheet.set_column(col_num, col_num, ancho)
        
        # Reiniciar para los nocturnos
        anchos_columnas = {}
        col_offset = 0
        
        # Escribir los dataframes nocturnos
        for i, fila in enumerate([9, 10, 11]):
            df_nocturno = dataframes[f'nocturno_fila{fila}']
            
            # Determinar qué columnas escribir
            if i == 0:  # Para el primer dataframe, incluir columna 'Nombre'
                worksheet.write(fila_inicio_nocturnos + 1, 0, "Nombre", formato_nombre_columna)
                ancho_titulo = len(df_nocturno.columns)
                col_offset_data = 1
            else:  # Para los otros dataframes, omitir columna 'Nombre'
                ancho_titulo = len(df_nocturno.columns) - 1
                col_offset_data = 0
            
            # Escribir el título
            worksheet.merge_range(
                fila_inicio_nocturnos, col_offset, 
                fila_inicio_nocturnos, col_offset + ancho_titulo - 1, 
                titulos_nocturnos[i], formato_titulo_nocturno
            )
            
            # Escribir los encabezados
            nombres_columnas = df_nocturno.columns[1:]  # Excluir 'Nombre'
            for j, col_name in enumerate(nombres_columnas):
                worksheet.write(fila_inicio_nocturnos + 1, col_offset + col_offset_data + j, 
                               col_name, formato_encabezado_nocturno)
            
            # Escribir los datos
            for row_idx in range(df_nocturno.shape[0]):
                # Escribir el nombre en la primera columna solo para el primer dataframe
                if i == 0:
                    nombre_estacion = df_nocturno.iloc[row_idx, 0]
                    worksheet.write(fila_inicio_nocturnos + row_idx + 2, 0, nombre_estacion, formato_emri)
                
                # Escribir los valores de las columnas de datos
                for j, col_idx in enumerate(range(1, df_nocturno.shape[1])):
                    valor = df_nocturno.iloc[row_idx, col_idx]
                    col_real = col_offset + col_offset_data + j
                    
                    # Si la columna es "Declaración", aplicar formato según el valor
                    if df_nocturno.columns[col_idx] == "Declaración":
                        if pd.notna(valor):
                            valor_lower = str(valor).lower()
                            if valor_lower == "pasa":
                                formato = formato_pasa
                            elif valor_lower == "no pasa":
                                formato = formato_no_pasa
                            elif "condicional" in valor_lower:
                                formato = formato_pasa_condicional
                            else:
                                formato = formato_datos
                            
                            worksheet.write(fila_inicio_nocturnos + row_idx + 2, col_real, valor, formato)
                    else:
                        # Para columnas numéricas
                        if pd.notna(valor):
                            if isinstance(valor, (int, float)):
                                worksheet.write_number(fila_inicio_nocturnos + row_idx + 2, col_real, valor, formato_numero)
                            else:
                                worksheet.write(fila_inicio_nocturnos + row_idx + 2, col_real, valor, formato_datos)
            
            # Actualizar el desplazamiento para el siguiente dataframe
            col_offset += ancho_titulo + 1
    
        # Ajustar automáticamente el ancho de las columnas
        worksheet.autofit()
    
    print(f"Todos los DataFrames guardados en formato especial en {ruta_excel_nuevo}")


def combinar_excels(archivo1, archivo2, archivo_salida):
    """
    Combina dos archivos Excel en uno nuevo.
    
    Args:
        archivo1 (str): Ruta al primer archivo Excel
        archivo2 (str): Ruta al segundo archivo Excel
        archivo_salida (str): Ruta donde guardar el archivo combinado
    """
    # Crear un nuevo libro de trabajo para el archivo combinado
    nuevo_libro = None
    
    # Procesar el primer archivo
    try:
        # Cargar el primer libro de trabajo
        libro1 = load_workbook(archivo1, data_only=False)
        
        # El primer libro será la base del nuevo libro combinado
        nuevo_libro = libro1
        print(f"Archivo {archivo1} procesado correctamente.")
    except Exception as e:
        print(f"Error al procesar {archivo1}: {str(e)}")
        return
    
    # Procesar el segundo archivo
    try:
        # Cargar el segundo libro de trabajo
        libro2 = load_workbook(archivo2, data_only=False)
        
        # Copiar cada hoja del segundo libro al nuevo libro
        for nombre_hoja in libro2.sheetnames:
            # Verificar si ya existe una hoja con ese nombre
            contador = 0
            nombre_final = nombre_hoja
            
            # Si ya existe una hoja con ese nombre, renombrarla
            while nombre_final in nuevo_libro.sheetnames:
                contador += 1
                nombre_final = f"{nombre_hoja} ({contador})"
            
            # Obtener la hoja del segundo libro
            hoja = libro2[nombre_hoja]
            
            # Crear una nueva hoja en el libro combinado
            nueva_hoja = nuevo_libro.create_sheet(title=nombre_final)
            
            # Copiar contenido celda por celda
            for row_idx in range(1, hoja.max_row + 1):
                for col_idx in range(1, hoja.max_column + 1):
                    # Obtener celda origen
                    celda_origen = hoja.cell(row=row_idx, column=col_idx)
                    
                    # Obtener celda destino
                    celda_destino = nueva_hoja.cell(row=row_idx, column=col_idx, value=celda_origen.value)
                    
                    # Intentar copiar el formato (ignora errores de estilo)
                    try:
                        if celda_origen.has_style:
                            # Usar un enfoque más seguro para copiar estilos
                            celda_destino.number_format = celda_origen.number_format
                            
                            # Intentar copiar otros atributos solo si existen
                            for attr in ['font', 'border', 'fill', 'alignment']:
                                try:
                                    setattr(celda_destino, attr, copy(getattr(celda_origen, attr)))
                                except:
                                    pass
                    except:
                        # Si hay algún error con los estilos, continuamos sin formato
                        pass
            
            # Copiar las celdas combinadas
            for rango_combinado in hoja.merged_cells.ranges:
                nueva_hoja.merge_cells(str(rango_combinado))
            
            # Intentar copiar anchos de columna
            try:
                for col_idx in range(1, hoja.max_column + 1):
                    column_letter = get_column_letter(col_idx)
                    if column_letter in hoja.column_dimensions:
                        if hoja.column_dimensions[column_letter].width is not None:
                            nueva_hoja.column_dimensions[column_letter].width = hoja.column_dimensions[column_letter].width
            except:
                pass
            
            # Intentar copiar altura de filas
            try:
                for row_idx in range(1, hoja.max_row + 1):
                    if row_idx in hoja.row_dimensions:
                        if hoja.row_dimensions[row_idx].height is not None:
                            nueva_hoja.row_dimensions[row_idx].height = hoja.row_dimensions[row_idx].height
            except:
                pass
        
        print(f"Archivo {archivo2} procesado correctamente.")
    except Exception as e:
        print(f"Error al procesar {archivo2}: {str(e)}")
    
    # Guardar el libro combinado
    try:
        nuevo_libro.save(archivo_salida)
        print(f"Archivo combinado guardado como {archivo_salida}")
    except Exception as e:
        print(f"Error al guardar el archivo combinado: {str(e)}")
