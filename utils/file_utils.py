import os
import re
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, PatternFill, Alignment, Protection
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def round_dataframe(df):
    """
    Redondea todos los valores numéricos en un DataFrame a 2 decimales
    
    Args:
        df: DataFrame a redondear
    
    Returns:
        DataFrame con valores redondeados
    """
    return df.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)

def is_merged_cell(sheet, row, col):
    """
    Verifica si una celda está combinada en una hoja de Excel
    
    Args:
        sheet: Hoja de Excel
        row: Número de fila
        col: Número de columna
    
    Returns:
        Booleano indicando si la celda está combinada
    """
    cell_ref = f"{get_column_letter(col)}{row}"
    for merged_range in sheet.merged_cells.ranges:
        if cell_ref in merged_range:
            return True
    return False

def combine_excel_files(carpeta):
    """
    Combina todos los archivos Excel de una carpeta en uno solo, 
    alternando hojas PTO y MET, y elimina los originales
    
    Args:
        carpeta: Ruta de la carpeta con los archivos Excel
    """
    # Crear un nuevo libro de Excel
    libro_destino = Workbook()
    libro_destino.remove(libro_destino.active)  # Elimina la hoja vacía por defecto

    # Diccionario para clasificar archivos por número
    archivos_por_grupo = {}

    # Recorrer los archivos en la carpeta y clasificarlos
    for archivo in os.listdir(carpeta):
        if archivo.endswith(".xlsx") and not archivo.startswith("~$"):  # Evita archivos temporales
            ruta_archivo = os.path.join(carpeta, archivo)

            # Extraer el número del archivo
            match = re.search(r'\d+', archivo)
            if not match:
                continue  # Si no tiene número, lo ignora
            
            numero_grupo = match.group()  # Extrae el número
            prefijo = "PTO" if "PTO" in archivo else "MET"  # Determina si es PTO o MET

            # Guardar el archivo en la lista correspondiente al grupo
            if numero_grupo not in archivos_por_grupo:
                archivos_por_grupo[numero_grupo] = {"PTO": None, "MET": None}
            archivos_por_grupo[numero_grupo][prefijo] = ruta_archivo

    # Procesar los archivos en orden intercalado (PTO primero, luego MET)
    for numero_grupo in sorted(archivos_por_grupo.keys(), key=int):
        for tipo in ["PTO", "MET"]:
            ruta_archivo = archivos_por_grupo[numero_grupo][tipo]
            if ruta_archivo:  # Verificar si el archivo existe
                nombre_hoja = f"{tipo}{numero_grupo}"  # Ejemplo: PTO1, MET1, PTO2, MET2
                
                # Cargar libro origen
                libro_origen = load_workbook(ruta_archivo)
                hoja_origen = libro_origen.active  # Solo toma la primera hoja

                # Crear hoja en el nuevo archivo
                hoja_nueva = libro_destino.create_sheet(title=nombre_hoja)

                # Copiar anchos de columnas
                for col in hoja_origen.column_dimensions:
                    hoja_nueva.column_dimensions[col].width = hoja_origen.column_dimensions[col].width

                # Copiar celdas combinadas
                merged_ranges = list(hoja_origen.merged_cells.ranges)
                for rango in merged_ranges:
                    hoja_nueva.merge_cells(start_row=rango.min_row, 
                                           end_row=rango.max_row,
                                           start_column=rango.min_col, 
                                           end_column=rango.max_col)

                # Copiar datos y estilos
                for fila in hoja_origen.iter_rows():
                    for celda in fila:
                        nueva_celda = hoja_nueva.cell(row=celda.row, column=celda.column, value=celda.value)

                        # Copiar estilos
                        if celda.has_style:
                            nueva_celda.font = Font(
                                name=celda.font.name, size=celda.font.size, bold=celda.font.bold,
                                italic=celda.font.italic, underline=celda.font.underline, color=celda.font.color
                            )
                            nueva_celda.border = Border(
                                left=celda.border.left, right=celda.border.right,
                                top=celda.border.top, bottom=celda.border.bottom
                            )
                            nueva_celda.fill = PatternFill(
                                fill_type=celda.fill.fill_type, start_color=celda.fill.start_color, end_color=celda.fill.end_color
                            )
                            nueva_celda.alignment = Alignment(
                                horizontal=celda.alignment.horizontal, vertical=celda.alignment.vertical, wrap_text=celda.alignment.wrap_text
                            )
                            nueva_celda.protection = Protection(locked=celda.protection.locked)

                libro_origen.close()

    # Guardar el archivo combinado
    ruta_salida = os.path.join(carpeta, "Excel_Intercalado.xlsx")
    libro_destino.save(ruta_salida)
    print(f"✅ Archivo combinado guardado en: {ruta_salida}")

    # Eliminar archivos originales
    for numero_grupo in archivos_por_grupo:
        for tipo in ["PTO", "MET"]:
            ruta_archivo = archivos_por_grupo[numero_grupo][tipo]
            if ruta_archivo and os.path.exists(ruta_archivo):  # Verifica que el archivo existe antes de eliminarlo
                os.remove(ruta_archivo)
                print(f"🗑️ Eliminado: {ruta_archivo}")

    print("✅ Todos los archivos originales han sido eliminados.")